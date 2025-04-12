from fastapi import FastAPI, HTTPException, Depends, Query, UploadFile, File, HTTPException
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, func, desc, and_
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from datetime import datetime, timedelta
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
from collections import defaultdict, Counter
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from io import BytesIO
from pydantic import BaseModel


DATABASE_URL = "sqlite:///./inventory.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- Models ---

class Inventory(Base):
    __tablename__ = "inventory"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True)
    quantity = Column(Float)
    unit = Column(String)
    price_per_unit = Column(Float)
    total_cost = Column(Float)
    type = Column(String)
    date_added = Column(DateTime, default=datetime.utcnow)

class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    item_name = Column(String)
    quantity = Column(Float)
    total_cost = Column(Float)
    date = Column(DateTime, default=datetime.utcnow)

class DishType(Base):
    __tablename__ = "dish_types"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True)

class Dish(Base):
    __tablename__ = "dishes"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True)
    type_id = Column(Integer, ForeignKey("dish_types.id"))
    type = relationship("DishType")

class DishIngredient(Base):
    __tablename__ = "dish_ingredients"
    id = Column(Integer, primary_key=True, index=True)
    ingredient_name = Column(String, index=True)  # <- ingredient name (e.g., "Tomato")
    dish_id = Column(Integer, ForeignKey("dishes.id"))
    quantity_required = Column(Float)
    dish = relationship("Dish")


class InventoryLog(Base):
    __tablename__ = "inventory_log"
    id = Column(Integer, primary_key=True, index=True)
    ingredient_id = Column(Integer, ForeignKey("inventory.id"))
    quantity_left = Column(Float)
    date = Column(DateTime, default=datetime.utcnow)
    ingredient = relationship("Inventory")

class IngredientInput(BaseModel):
    name: str
    quantity_required: float

class AddDishRequest(BaseModel):
    name: str
    type: str
    ingredients: List[IngredientInput]

class DishIngredientOut(BaseModel):
    ingredient_name: str
    quantity_required: float

class DishOut(BaseModel):
    id: int
    name: str
    type: str
    ingredients: List[DishIngredientOut]

    class Config:
        orm_mode = True

class DishIngredientUpdate(BaseModel):
    ingredient_name: str
    quantity_required: float

class DishUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    ingredients: Optional[List[DishIngredientUpdate]] = None

class PrepareDishRequest(BaseModel):
    dish_name: str
    quantity: float
    date: Optional[str] = None  # format: YYYY-MM-DD


Base.metadata.create_all(bind=engine)

# --- Routes ---

@app.post("/add_item")
def add_item(
    name: str,
    quantity: float,
    unit: str,
    price_per_unit: Optional[float] = None,
    total_cost: Optional[float] = None,
    type: Optional[str] = None,
    date_added: datetime = datetime.utcnow(),
    db: Session = Depends(get_db)
):

    if price_per_unit is not None:
        total_cost = quantity * price_per_unit
    elif total_cost is not None:
        total_cost = total_cost
    else:
        raise HTTPException(status_code=400, detail="Either price_per_unit or total_cost must be provided")


    item = Inventory(
        name=name,
        quantity=quantity,
        unit=unit,
        price_per_unit=price_per_unit if price_per_unit is not None else 0.0,
        total_cost=total_cost,
        type=type if type is not None else "",
        date_added=date_added
    )

    db.add(item)
    db.add(Expense(item_name=name, quantity=quantity, total_cost=total_cost, date=date_added))
    db.commit()
    db.refresh(item)

    inventory = db.query(Inventory).all()
    return {"message": "Item added successfully!", "inventory": inventory}

@app.get("/search_inventory")
def search_inventory(
    name: Optional[str] = Query(None, description="Partial or full item name"),
    type: Optional[str] = Query(None, description="Inventory type (e.g. 'Vegetables', 'Dairy')"),
    start_date: Optional[str] = Query(None, description="Start date in YYYY-MM-DD format"),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Inventory)

        if name:
            query = query.filter(Inventory.name.ilike(f"%{name}%"))
        elif type:
            query = query.filter(Inventory.type.ilike(f"%{type}%"))

        if start_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(Inventory.date_added >= start)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD.")

        if end_date:
            try:
                end = datetime.strptime(end_date, "%Y-%m-%d")
                query = query.filter(Inventory.date_added <= end)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD.")

        results = query.all()

        return [
            {
                "id": item.id,
                "name": item.name,
                "price_per_unit": item.price_per_unit,
                "unit": item.unit,
                "quantity": item.quantity,
                "total_cost": item.total_cost,
                "type": item.type,
                "date_added": item.date_added.strftime("%Y-%m-%d %H:%M:%S")
            }
            for item in results
        ]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")

@app.delete("/delete_item/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    db.delete(item)
    db.commit()
    return {
        "message": "Item deleted successfully!",
        "current_inventory": db.query(Inventory).all()
    }


@app.put("/update_item/{item_id}")
def update_item(
    item_id: int,
    name: str = Query(...),
    quantity: float = Query(...),
    unit: str = Query(...),
    price_per_unit: float = Query(None),
    total_cost: float = Query(None),
    date_added: Optional[datetime] = Query(default=datetime.utcnow()),
    type: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Recalculate total_cost
    if price_per_unit is not None:
        total_cost = quantity * price_per_unit
    elif total_cost is not None:
        total_cost = total_cost
    else:
        raise HTTPException(status_code=400, detail="Either price_per_unit or total_cost must be provided")

    # Update fields
    item.name = name
    item.quantity = quantity
    item.unit = unit
    item.price_per_unit = price_per_unit if price_per_unit is not None else 0.0
    item.total_cost = total_cost
    item.date_added = date_added
    item.type = type if type is not None else ""

    db.commit()
    db.refresh(item)

    return {
        "message": "Item updated successfully!",
        "updated_item": {
            "id": item.id,
            "name": item.name,
            "quantity": item.quantity,
            "unit": item.unit,
            "price_per_unit": item.price_per_unit,
            "total_cost": item.total_cost,
            "type": item.type,
            "date_added": item.date_added.strftime("%Y-%m-%d")
        },
        "current_inventory": [
            {
                "id": i.id,
                "name": i.name,
                "quantity": i.quantity,
                "unit": i.unit,
                "price_per_unit": i.price_per_unit,
                "total_cost": i.total_cost,
                "type": i.type,
                "date_added": i.date_added.strftime("%Y-%m-%d")
            } for i in db.query(Inventory).all()
        ]
    }

@app.get("/inventory")
def get_inventory(db: Session = Depends(get_db)):
    inventory_items = db.query(Inventory).all()
    return [
        {
            "id": item.id,
            "name": item.name,
            "price_per_unit": item.price_per_unit,
            "unit": item.unit,
            "quantity": item.quantity,
            "total_cost": item.total_cost,
            "type": item.type,
            "date_added": item.date_added.strftime("%Y-%m-%d %H:%M:%S")
        }
        for item in inventory_items
    ]

@app.get("/inventory_by_name/{item_name}")
def get_inventory_by_name(item_name: str, db: Session = Depends(get_db)):
    items = db.query(Inventory).filter(Inventory.name.ilike(f"%{item_name}%")).all()
    if not items:
        raise HTTPException(status_code=404, detail="Item not found")
    return items

@app.delete("/delete_all_inventory")
def delete_all_inventory(confirm: bool = False, db: Session = Depends(get_db)):
    if not confirm:
        raise HTTPException(status_code=400, detail="Please confirm deletion by setting confirm=true")

    deleted = db.query(Inventory).delete()
    db.commit()
    return {
        "message": f"Deleted {deleted} item(s) from inventory.",
        "current_inventory": db.query(Inventory).all()
    }


@app.get("/expense_report")
def expense_report(
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    inventory_name: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    # Determine date range if not provided
    if not start_date or not end_date:
        date_range = db.query(
            func.min(Inventory.date_added),
            func.max(Inventory.date_added)
        ).first()
        if not date_range or not date_range[0] or not date_range[1]:
            return {
                "message": "No inventory records in the database.",
                "total_expense": 0,
                "average_expense": 0,
                "highest_expense_day": None,
                "lowest_expense_day": None,
                "highest_expense_item": None,
                "lowest_expense_item": None,
                "most_frequent_inventory": None
            }
        start = date_range[0]
        end = date_range[1]
    else:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")

    # Base query
    query = db.query(Inventory).filter(
        Inventory.date_added >= start,
        Inventory.date_added <= end
    )

    # Apply inventory_name filter if present
    if inventory_name:
        query = query.filter(Inventory.name.ilike(f"%{inventory_name}%"))
    # Else apply type filter if present
    elif type:
        query = query.filter(Inventory.type.ilike(f"%{type}%"))

    inventory_items = query.all()

    if not inventory_items:
        return {
            "message": "No inventory found for the given filters.",
            "total_expense": 0,
            "average_expense": 0,
            "highest_expense_day": None,
            "lowest_expense_day": None,
            "highest_expense_item": None,
            "lowest_expense_item": None,
            "most_frequent_inventory": None
        }

    # Use total_cost directly
    total_expense = sum(item.total_cost for item in inventory_items)
    average_expense = total_expense / len(inventory_items)

    # Daily totals
    daily_expenses = defaultdict(float)
    for item in inventory_items:
        daily_expenses[item.date_added.date()] += item.total_cost

    highest_day = max(daily_expenses.items(), key=lambda x: x[1])
    lowest_day = min(daily_expenses.items(), key=lambda x: x[1])

    highest_expense_item = None
    lowest_expense_item = None
    most_frequent_inventory = None

    # Apply extended analysis only if inventory_name is NOT provided
    if not inventory_name:
        highest_item = max(inventory_items, key=lambda x: x.total_cost, default=None)
        lowest_item = min(inventory_items, key=lambda x: x.total_cost, default=None)
        highest_expense_item = highest_item.name if highest_item else None
        lowest_expense_item = lowest_item.name if lowest_item else None

        item_counts = Counter(item.name for item in inventory_items)
        most_frequent_inventory = item_counts.most_common(1)[0][0] if item_counts else None

    return {
        "inventory_name": inventory_name,
        "type": type if not inventory_name else None,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "total_expense": total_expense,
        "average_expense": average_expense,
        "highest_expense_day": {
            "date": highest_day[0].isoformat(),
            "amount": highest_day[1]
        },
        "lowest_expense_day": {
            "date": lowest_day[0].isoformat(),
            "amount": lowest_day[1]
        },
        "highest_expense_item": highest_expense_item,
        "lowest_expense_item": lowest_expense_item,
        "most_frequent_inventory": most_frequent_inventory
    }

@app.post("/upload_inventory_excel")
async def upload_inventory_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file (.xlsx or .xls)")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        required_columns = {"name", "quantity", "unit", "date_added"}
        optional_columns = {"price_per_unit", "total_cost", "type"}

        missing_required = required_columns - set(headers)
        if missing_required:
            return JSONResponse(status_code=400, content={
                "error": f"Missing required columns: {', '.join(missing_required)}. Found: {headers}"
            })

        col_index = {key: headers.index(key) for key in required_columns.union(optional_columns) if key in headers}

        added_items = []
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                name = row[col_index["name"]]
                quantity = float(row[col_index["quantity"]])
                unit = row[col_index["unit"]]

                price_per_unit = float(row[col_index["price_per_unit"]]) if "price_per_unit" in col_index and row[col_index["price_per_unit"]] is not None else None
                total_cost = float(row[col_index["total_cost"]]) if "total_cost" in col_index and row[col_index["total_cost"]] is not None else None
                type_ = str(row[col_index["type"]]).strip() if "type" in col_index and row[col_index["type"]] is not None else ""

                if price_per_unit is not None:
                    total_cost = quantity * price_per_unit
                elif total_cost is not None:
                    price_per_unit = total_cost / quantity
                else:
                    skipped_rows.append(f"Row {idx}: Missing price_per_unit and total_cost.")
                    continue

                date_raw = row[col_index["date_added"]]
                if isinstance(date_raw, datetime):
                    date_added = date_raw
                elif isinstance(date_raw, str):
                    date_added = datetime.strptime(date_raw.strip(), "%Y-%m-%d")
                else:
                    raise ValueError("Invalid date format in 'date_added'")

                # Check if an exact item already exists
                existing = db.query(Inventory).filter(
                    Inventory.name == name,
                    Inventory.quantity == quantity,
                    Inventory.unit == unit,
                    Inventory.price_per_unit == price_per_unit,
                    Inventory.total_cost == total_cost,
                    Inventory.type == type_,
                    Inventory.date_added == date_added
                ).first()

                if existing:
                    skipped_rows.append(f"Row {idx}: Exact item already exists. Skipped.")
                    continue

                item = Inventory(
                    name=name,
                    quantity=quantity,
                    unit=unit,
                    price_per_unit=price_per_unit,
                    total_cost=total_cost,
                    type=type_,
                    date_added=date_added
                )
                db.add(item)
                db.add(Expense(item_name=name, quantity=quantity, total_cost=total_cost, date=date_added))
                added_items.append(name)

            except Exception as row_error:
                skipped_rows.append(f"Row {idx}: {str(row_error)}")

        db.commit()

        return {
            "message": "Excel processed successfully",
            "added_items": added_items,
            "skipped_rows": skipped_rows
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@app.post("/upload_dish_excel")
async def upload_dish_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file (.xlsx or .xls)")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        required_columns = {"name", "type", "ingredient_name", "quantity_required"}

        missing_required = required_columns - set(headers)
        if missing_required:
            return JSONResponse(status_code=400, content={
                "error": f"Missing required columns: {', '.join(missing_required)}. Found: {headers}"
            })

        col_index = {key: headers.index(key) for key in required_columns}

        dishes_map = {}  # Temporary cache to hold dishes before commit
        added_dishes = []
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                dish_name = str(row[col_index["name"]]).strip()
                dish_type = str(row[col_index["type"]]).strip()
                ingredient_name = str(row[col_index["ingredient_name"]]).strip()
                quantity_required = float(row[col_index["quantity_required"]])

                # Fetch or create DishType
                dish_type_obj = db.query(DishType).filter_by(name=dish_type).first()
                if not dish_type_obj:
                    dish_type_obj = DishType(name=dish_type)
                    db.add(dish_type_obj)
                    db.flush()  # To assign an ID

                # Unique key for dish mapping
                dish_key = (dish_name, dish_type_obj.id)
                if dish_key not in dishes_map:
                    dish = db.query(Dish).filter_by(name=dish_name, type_id=dish_type_obj.id).first()
                    if not dish:
                        dish = Dish(name=dish_name, type_id=dish_type_obj.id)
                        db.add(dish)
                        db.flush()  # Assign ID
                        added_dishes.append(dish_name)
                    dishes_map[dish_key] = dish
                else:
                    dish = dishes_map[dish_key]

                # Add ingredient
                dish_ingredient = DishIngredient(
                    dish_id=dish.id,
                    ingredient_name=ingredient_name,
                    quantity_required=quantity_required
                )
                db.add(dish_ingredient)

            except Exception as row_error:
                skipped_rows.append(f"Row {idx}: {str(row_error)}")

        db.commit()

        return {
            "message": "Dishes uploaded successfully",
            "added_dishes": list(set(added_dishes)),
            "skipped_rows": skipped_rows
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@app.post("/add_dish")
def add_dish(request: AddDishRequest, db: Session = Depends(get_db)):
    # Check if dish with same name exists
    existing_dish = db.query(Dish).filter(Dish.name == request.name).first()
    if existing_dish:
        raise HTTPException(status_code=400, detail="Dish with this name already exists.")

    # Get or create DishType
    dish_type = db.query(DishType).filter(DishType.name == request.type).first()
    if not dish_type:
        dish_type = DishType(name=request.type)
        db.add(dish_type)
        db.commit()
        db.refresh(dish_type)

    # Create new dish
    new_dish = Dish(name=request.name, type_id=dish_type.id)
    db.add(new_dish)
    db.commit()
    db.refresh(new_dish)

    # Process ingredients
    for ing in request.ingredients:
        #inventory_item = db.query(Inventory).filter(Inventory.name == ing.name).first()
        #if not inventory_item:
            #raise HTTPException(status_code=404, detail=f"Ingredient '{ing.name}' not found in inventory.")

        dish_ingredient = DishIngredient(
            dish_id=new_dish.id,
            #ingredient_name=inventory_item.name,
            ingredient_name=ing.name,
            quantity_required=ing.quantity_required
        )
        db.add(dish_ingredient)

    db.commit()
    return {"message": f"Dish '{request.name}' added successfully with ingredients."}

@app.get("/dishes", response_model=List[DishOut])
def list_dishes(db: Session = Depends(get_db)):
    dishes = db.query(Dish).all()
    result = []

    for dish in dishes:
        # Get the dish type name
        dish_type = db.query(DishType).filter(DishType.id == dish.type_id).first()

        # Get all ingredients for this dish
        ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

        # Convert DishIngredient DB entries to response models
        ingredient_list = [
            DishIngredientOut(
                ingredient_name=di.ingredient_name,
                quantity_required=di.quantity_required
            ) for di in ingredients if di.ingredient_name is not None
        ]

        # Append formatted dish to result
        result.append(DishOut(
            id=dish.id,
            name=dish.name,
            type=dish_type.name if dish_type else "Unknown",
            ingredients=ingredient_list
        ))

    return result


@app.get("/dishes/by_name", response_model=List[DishOut])
def search_dishes_by_name(partial_name: str, db: Session = Depends(get_db)):
    matched_dishes = db.query(Dish).filter(
        Dish.name.ilike(f"%{partial_name}%")
    ).all()

    if not matched_dishes:
        raise HTTPException(status_code=404, detail="No matching dishes found")

    result = []
    for dish in matched_dishes:
        dish_type = db.query(DishType).filter(DishType.id == dish.type_id).first()
        ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

        ingredient_list = [
            DishIngredientOut(
                ingredient_name=di.ingredient_name,
                quantity_required=di.quantity_required
            )
            for di in ingredients
        ]

        result.append(DishOut(
            id=dish.id,
            name=dish.name,
            type=dish_type.name if dish_type else "Unknown",
            ingredients=ingredient_list
        ))

    return result

@app.delete("/dishes/{dish_name}")
def delete_dish_by_name(
    dish_name: str,
    confirm: bool = Query(False, description="Set to true to confirm deletion"),
    db: Session = Depends(get_db)
):
    dish = db.query(Dish).filter(Dish.name.ilike(dish_name)).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    if not confirm:
        raise HTTPException(
            status_code=400,
            detail=f"Deletion not confirmed. To delete dish '{dish.name}', set confirm=true."
        )

    db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).delete()
    db.delete(dish)
    db.commit()
    return {"message": f"Dish '{dish.name}' deleted successfully"}

@app.get("/dish_types")
def get_dish_types(db: Session = Depends(get_db)):
    dish_types = db.query(DishType).all()
    return [dt.name for dt in dish_types]


@app.get("/dishes/{dish_id}/cost")
def get_dish_cost(dish_id: int, db: Session = Depends(get_db)):
    dish = db.query(Dish).filter(Dish.id == dish_id).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()
    total_cost = 0.0

    for di in ingredients:
        inventory_item = db.query(Inventory).filter(Inventory.name == di.ingredient_name).order_by(Inventory.date).first()
        if not inventory_item:
            raise HTTPException(status_code=400, detail=f"Ingredient '{di.ingredient_name}' not found in inventory")
        item_cost = di.quantity_required * inventory_item.unit_price
        total_cost += item_cost

    return {
        "dish_id": dish.id,
        "dish_name": dish.name,
        "total_cost": round(total_cost, 2)
    }

@app.put("/dishes/{dish_id}")
def update_dish(dish_id: int, payload: DishUpdate, db: Session = Depends(get_db)):
    dish = db.query(Dish).filter(Dish.id == dish_id).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    # Update dish name and type
    dish.name = payload.name

    # Lookup dish type
    dish_type = db.query(DishType).filter(DishType.name.ilike(payload.type)).first()
    if not dish_type:
        dish_type = DishType(name=payload.type)
        db.add(dish_type)
        db.commit()
        db.refresh(dish_type)

    dish.type_id = dish_type.id
    db.commit()

    dish_ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

    # Update ingredients: map by name for easy comparison
    existing_ingredients = {di.ingredient_name.lower(): di for di in dish_ingredients}
    updated_names = {ing.ingredient_name.lower() for ing in payload.ingredients}

    for ing in payload.ingredients:
        key = ing.ingredient_name.lower()
        if key in existing_ingredients:
            existing_ingredients[key].quantity_required = ing.quantity_required
        else:
            db.add(DishIngredient(
                dish_id=dish.id,
                ingredient_name=ing.ingredient_name,
                quantity_required=ing.quantity_required
            ))

    # Delete ingredients no longer in request

    for key, di in existing_ingredients.items():
        if key not in updated_names:
            db.delete(di)

    db.commit()
    return {"message": "Dish updated successfully"}


@app.post("/prepare_dish")
def prepare_dish(
    request: PrepareDishRequest,
    db: Session = Depends(get_db)
):
    # Parse the preparation date (or default to now)
    if request.date:
        try:
            prepare_date = datetime.strptime(request.date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    else:
        prepare_date = datetime.utcnow()

    dish = db.query(Dish).filter(Dish.name.ilike(request.dish_name.strip())).first()
    if not dish:
        raise HTTPException(status_code=404, detail=f"Dish '{request.dish_name}' not found")

    dish_ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()
    if not dish_ingredients:
        raise HTTPException(status_code=400, detail=f"No ingredients found for dish '{request.dish_name}'")

    usage_summary = []

    for ingredient in dish_ingredients:
        required_qty = ingredient.quantity_required * request.quantity

        inventory_batches = db.query(Inventory).filter(
            Inventory.name.ilike(ingredient.ingredient_name),
            Inventory.quantity > 0
        ).order_by(Inventory.date_added.asc()).all()

        if not inventory_batches:
            raise HTTPException(
                status_code=400,
                detail=f"No inventory available for {ingredient.ingredient_name}"
            )

        for batch in inventory_batches:
            if required_qty <= 0:
                break

            unit = batch.unit.strip().lower()
            batch_qty_in_base = batch.quantity

            # Convert to base unit (g, ml, pieces)
            if unit == "kg":
                batch_qty_in_base *= 1000
            elif unit in ["g", "gm", "grams"]:
                pass
            elif unit in ["liter", "litre", "l"]:
                batch_qty_in_base *= 1000
            elif unit == "ml":
                pass
            elif unit in ["piece", "pieces", "pc", "pcs", "pack", "packs"]:
                pass
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported unit '{batch.unit}' for ingredient '{batch.name}'")

            deduct_qty = min(batch_qty_in_base, required_qty)
            remaining_qty_in_base = batch_qty_in_base - deduct_qty
            required_qty -= deduct_qty

            # Convert back to original unit
            if unit == "kg":
                batch.quantity = remaining_qty_in_base / 1000
            elif unit in ["g", "gm", "grams"]:
                batch.quantity = remaining_qty_in_base
            elif unit in ["liter", "litre", "l"]:
                batch.quantity = remaining_qty_in_base / 1000
            elif unit == "ml":
                batch.quantity = remaining_qty_in_base
            elif unit in ["piece", "pieces", "pc", "pcs", "pack", "packs"]:
                batch.quantity = remaining_qty_in_base

            db.add(batch)

            # Insert log at specified date
            log_entry = InventoryLog(
                ingredient_id=batch.id,
                quantity_left=batch.quantity,
                date=prepare_date
            )
            db.add(log_entry)

            usage_summary.append({
                "ingredient": ingredient.ingredient_name,
                "inventory_batch_id": batch.id,
                "used_from_batch": deduct_qty,
                "remaining_in_batch": batch.quantity,
                "unit": unit,
                "logged_at": prepare_date.strftime("%Y-%m-%d")
            })

            # Update all future logs
            future_logs = db.query(InventoryLog).filter(
                and_(
                    InventoryLog.ingredient_id == batch.id,
                    InventoryLog.date > prepare_date
                )
            ).order_by(InventoryLog.date.asc()).all()

            current_quantity = batch.quantity

            for log in future_logs:
                log.quantity_left = current_quantity
                db.add(log)

            # ✅ Update Inventory table based on the latest log
            latest_log = db.query(InventoryLog).filter(
                InventoryLog.ingredient_id == batch.id
            ).order_by(InventoryLog.date.desc()).first()

            if latest_log:
                inventory_entry = db.query(Inventory).filter(Inventory.id == batch.id).first()
                if inventory_entry:
                    inventory_entry.quantity = latest_log.quantity_left
                    db.add(inventory_entry)

        if required_qty > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Not enough '{ingredient.ingredient_name}' in inventory to prepare {request.quantity} servings. Short by {required_qty:.2f} units."
            )

    db.commit()

    return {
        "message": f"Dish '{dish.name}' prepared successfully for {request.quantity} servings on {prepare_date.date()}.",
        "usage_summary": usage_summary
    }


@app.get("/inventory_on_date")
def inventory_on_date(date: str, db: Session = Depends(get_db)):
    try:
        date_parsed = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    next_day = date_parsed + timedelta(days=1)

    # Get all logs for the day
    logs = db.query(InventoryLog).filter(
        InventoryLog.date >= date_parsed,
        InventoryLog.date < next_day
    ).order_by(InventoryLog.ingredient_id, desc(InventoryLog.date)).all()

    # Track the latest log per ingredient_id
    latest_log_map = {}
    for log in logs:
        if log.ingredient_id not in latest_log_map:
            latest_log_map[log.ingredient_id] = log  # first (latest) log due to descending sort

    # Convert to list for response
    response = []
    for log in latest_log_map.values():
        # Optional: join with inventory to get ingredient name
        inventory_item = db.query(Inventory).filter(Inventory.id == log.ingredient_id).first()
        response.append({
            "ingredient_id": log.ingredient_id,
            "inventory_name": inventory_item.name if inventory_item else "Unknown",
            "unit": inventory_item.unit if inventory_item else "",
            "quantity_left": log.quantity_left,
            "log_time": log.date.strftime("%Y-%m-%d %H:%M:%S")
        })

    return response






if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
