from fastapi import FastAPI, HTTPException, Depends, Query, UploadFile, File, HTTPException
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, func, desc, and_, text, \
    inspect
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from datetime import datetime, timedelta, date
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
from collections import defaultdict, Counter
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from io import BytesIO
from pydantic import BaseModel
from dateutil import parser
import openai
import os
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# from your_existing_models import Inventory, Expense, Dish, DishIngredient, get_db

# Ideally, set this as an environment variable in production
# openai.api_key = os.getenv("OPENAI_API_KEY", "")
#DATABASE_URL = "sqlite:///./inventory.db"
#engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./inventory.db")

if DATABASE_URL.startswith("postgresql://") or DATABASE_URL.startswith("postgres://"):
    # Fix postgres:// URL (Heroku/Render often use this)
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

    engine = create_engine(DATABASE_URL)
    print("Using PostgreSQL database")
else:
    # SQLite for local development
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
    print("Using SQLite database")

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# --- Migration Handler ---
class RenderSafeMigration:
    def __init__(self):
        self.database_url = DATABASE_URL
        self.engine = engine

    def check_schema_on_startup(self):
        """Check database schema on startup but don't auto-migrate"""
        try:
            logger.info("Checking database schema on startup...")

            inspector = inspect(self.engine)

            # Check if dish_ingredients table exists
            if 'dish_ingredients' not in inspector.get_table_names():
                logger.info("dish_ingredients table not found. Creating tables...")
                Base.metadata.create_all(bind=self.engine)
                return True

            # Check if unit column exists
            columns = inspector.get_columns('dish_ingredients')
            column_names = [col['name'] for col in columns]

            if 'unit' not in column_names:
                logger.warning("Unit column not found in dish_ingredients table.")
                logger.info("You can add it manually using the migration endpoint.")
                return True  # Don't fail startup
            else:
                logger.info("Database schema is up to date.")
                return True

        except Exception as e:
            logger.error(f"Schema check failed: {e}")
            return False

    def add_unit_column(self):
        """Add unit column with simple default value"""
        try:
            with self.engine.connect() as connection:
                trans = connection.begin()

                try:
                    # Add column with default
                    logger.info("Adding unit column to dish_ingredients...")
                    connection.execute(text(
                        "ALTER TABLE dish_ingredients ADD COLUMN unit VARCHAR DEFAULT 'gm'"
                    ))

                    # Update existing records to use default 'gm'
                    logger.info("Setting default unit 'gm' for existing records...")
                    connection.execute(text("""
                        UPDATE dish_ingredients 
                        SET unit = 'gm'
                        WHERE unit IS NULL OR unit = ''
                    """))

                    # Verify the migration
                    result = connection.execute(text(
                        "SELECT COUNT(*) FROM dish_ingredients WHERE unit IS NOT NULL"
                    ))
                    updated_count = result.scalar()

                    trans.commit()
                    logger.info(f"Migration completed successfully. Updated {updated_count} records.")
                    return True

                except Exception as e:
                    logger.error(f"Migration failed: {e}")
                    trans.rollback()
                    return False

        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            return False

    def add_costing_column(self):
        """Add cost_per_unit column with intelligent default calculation"""
        try:
            with self.engine.connect() as connection:
                trans = connection.begin()

                try:
                    # Check if column already exists
                    inspector = inspect(self.engine)
                    columns = [col['name'] for col in inspector.get_columns('dish_ingredients')]

                    if 'cost_per_unit' in columns:
                        logger.info("Cost_per_unit column already exists.")
                        trans.rollback()
                        return True

                    # Add column with default
                    logger.info("Adding cost_per_unit column to dish_ingredients...")
                    connection.execute(text(
                        "ALTER TABLE dish_ingredients ADD COLUMN cost_per_unit REAL DEFAULT 0.0"
                    ))

                    # Calculate and update cost for existing records
                    logger.info("Calculating costs for existing ingredients...")

                    # Get all existing dish ingredients
                    existing_ingredients = connection.execute(text("""
                        SELECT di.id, di.ingredient_name, di.quantity_required, di.unit
                        FROM dish_ingredients di
                        WHERE di.cost_per_unit IS NULL OR di.cost_per_unit = 0.0
                    """)).fetchall()

                    updated_count = 0
                    for ingredient in existing_ingredients:
                        ingredient_id, ingredient_name, quantity_required, unit = ingredient

                        # Find the most recent inventory item for this ingredient
                        inventory_result = connection.execute(text("""
                            SELECT price_per_unit, unit as inv_unit
                            FROM inventory 
                            WHERE LOWER(name) LIKE LOWER(:ingredient_name)
                            ORDER BY date_added DESC 
                            LIMIT 1
                        """), {"ingredient_name": f"%{ingredient_name}%"}).fetchone()

                        if inventory_result:
                            price_per_unit, inv_unit = inventory_result

                            # Calculate cost considering unit conversion if needed
                            cost_per_unit = self._calculate_cost_with_unit_conversion(
                                price_per_unit, inv_unit, unit or 'gm'
                            )

                            # Update the ingredient cost
                            connection.execute(text("""
                                UPDATE dish_ingredients 
                                SET cost_per_unit = :cost_per_unit
                                WHERE id = :ingredient_id
                            """), {
                                "cost_per_unit": cost_per_unit,
                                "ingredient_id": ingredient_id
                            })

                            updated_count += 1
                        else:
                            # No inventory found, set default cost
                            logger.warning(f"No inventory found for ingredient: {ingredient_name}")
                            connection.execute(text("""
                                UPDATE dish_ingredients 
                                SET cost_per_unit = 1.0
                                WHERE id = :ingredient_id
                            """), {"ingredient_id": ingredient_id})

                    trans.commit()
                    logger.info(
                        f"Costing column migration completed successfully. Updated {updated_count} records with calculated costs.")
                    return True

                except Exception as e:
                    logger.error(f"Costing column migration failed: {e}")
                    trans.rollback()
                    return False

        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            return False


# Initialize migration handler
migration_handler = RenderSafeMigration()


# --- Models ---

class Inventory(Base):
    __tablename__ = "inventory"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True)
    quantity = Column(Float)
    unit = Column(String)
    price_per_unit = Column(Float)
    total_cost = Column(Float)
    type = Column(String)
    date_added = Column(DateTime, default=datetime.utcnow)


class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    item_name = Column(String)
    quantity = Column(Float)
    total_cost = Column(Float)
    date = Column(DateTime, default=datetime.utcnow)


class DishType(Base):
    __tablename__ = "dish_types"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True)


class Dish(Base):
    __tablename__ = "dishes"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True)
    type_id = Column(Integer, ForeignKey("dish_types.id"))
    type = relationship("DishType")


class DishIngredient(Base):
    __tablename__ = "dish_ingredients"
    id = Column(Integer, primary_key=True, index=True)
    ingredient_name = Column(String, index=True)  # <- ingredient name (e.g., "Tomato")
    dish_id = Column(Integer, ForeignKey("dishes.id"))
    quantity_required = Column(Float)
    unit = Column(String, default="gm")  # NEW COLUMN with default
    dish = relationship("Dish")


class InventoryLog(Base):
    __tablename__ = "inventory_log"
    id = Column(Integer, primary_key=True, index=True)
    ingredient_id = Column(Integer, ForeignKey("inventory.id"))
    quantity_left = Column(Float)
    date = Column(DateTime, default=datetime.utcnow)
    ingredient = relationship("Inventory")


class IngredientInput(BaseModel):
    name: str
    quantity_required: float
    unit: str = "gm"  # Default unit


class AddDishRequest(BaseModel):
    name: str
    type: str
    ingredients: List[IngredientInput]


class DishIngredientOut(BaseModel):
    ingredient_name: str
    quantity_required: float
    unit: str = "gm"  # Default for backward compatibility


class DishOut(BaseModel):
    id: int
    name: str
    type: str
    ingredients: List[DishIngredientOut]

    class Config:
        orm_mode = True


class DishIngredientUpdate(BaseModel):
    ingredient_name: str
    quantity_required: float
    unit: str = "gm"  # Default unit


class DishUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    ingredients: Optional[List[DishIngredientUpdate]] = None


class PrepareDishRequest(BaseModel):
    dish_name: str
    quantity: float
    date: Optional[str] = None  # format: YYYY-MM-DD


class OpenAIPromptRequest(BaseModel):
    prompt: str


# Database initialization
def initialize_database():
    """Initialize database with schema check only"""
    try:
        # First, create basic tables if they don't exist
        Base.metadata.create_all(bind=engine)

        # Then run schema check (no auto-migration)
        schema_ok = migration_handler.check_schema_on_startup()

        if not schema_ok:
            logger.error("Database schema check failed!")

        return schema_ok

    except Exception as e:
        logger.error(f"Database initialization failed: {e}")
        return False


# Call this at startup
database_ready = initialize_database()


# --- Migration Endpoints ---
@app.post("/admin/migrate-add-unit-column")
def manual_add_unit_column(
        confirm: bool = Query(False, description="Set to true to confirm migration"),
        db: Session = Depends(get_db)
):
    """Manual endpoint to add unit column - you control when to run this"""
    if not confirm:
        return {
            "message": "Migration not confirmed. Set confirm=true to proceed.",
            "warning": "This will modify the dish_ingredients table structure.",
            "current_status": get_migration_status(db)
        }

    try:
        success = migration_handler.add_unit_column()
        if success:
            return {
                "message": "Unit column added successfully! All existing ingredients set to 'gm'.",
                "status": get_migration_status(db)
            }
        else:
            return {
                "message": "Migration failed. Check logs for details.",
                "status": "failed"
            }
    except Exception as e:
        return {
            "message": f"Migration error: {str(e)}",
            "status": "error"
        }


@app.get("/system/migration-status")
def get_migration_status(db: Session = Depends(get_db)):
    """Check migration status"""
    try:
        inspector = inspect(engine)
        columns = [col['name'] for col in inspector.get_columns('dish_ingredients')]
        has_unit_column = 'unit' in columns

        # Count records with/without units
        total_ingredients = db.query(DishIngredient).count()
        ingredients_with_units = db.query(DishIngredient).filter(
            DishIngredient.unit.isnot(None),
            DishIngredient.unit != ""
        ).count()

        return {
            "migration_complete": has_unit_column and (ingredients_with_units == total_ingredients),
            "schema_updated": has_unit_column,
            "data_migrated": ingredients_with_units == total_ingredients,
            "stats": {
                "total_ingredients": total_ingredients,
                "ingredients_with_units": ingredients_with_units,
                "completion_percentage": round((ingredients_with_units / total_ingredients * 100),
                                               2) if total_ingredients > 0 else 100
            }
        }

    except Exception as e:
        return {
            "migration_complete": False,
            "error": str(e)
        }


@app.post("/admin/migrate-add-costing-column")
def manual_add_costing_column(
        confirm: bool = Query(False, description="Set to true to confirm migration"),
        db: Session = Depends(get_db)
):
    """Manual endpoint to add cost_per_unit column"""
    if not confirm:
        return {
            "message": "Migration not confirmed. Set confirm=true to proceed.",
            "warning": "This will modify the dish_ingredients table structure and calculate costs.",
            "current_status": get_migration_status(db)
        }

    try:
        success = migration_handler.add_costing_column()
        if success:
            return {
                "message": "Costing column added successfully! Costs calculated from inventory prices.",
                "status": get_migration_status(db)
            }
        else:
            return {
                "message": "Migration failed. Check logs for details.",
                "status": "failed"
            }
    except Exception as e:
        return {
            "message": f"Migration error: {str(e)}",
            "status": "error"
        }


@app.post("/admin/migrate-all")
def run_all_migrations(
        confirm: bool = Query(False, description="Set to true to confirm all migrations"),
        db: Session = Depends(get_db)
):
    """Run all pending migrations at once"""
    if not confirm:
        return {
            "message": "Migration not confirmed. Set confirm=true to proceed.",
            "warning": "This will run all pending migrations on dish_ingredients table.",
            "current_status": get_migration_status(db)
        }

    try:
        success, migrations_run = migration_handler.run_all_migrations()
        if success:
            return {
                "message": f"All migrations completed successfully! Ran: {', '.join(migrations_run) if migrations_run else 'none needed'}",
                "migrations_run": migrations_run,
                "status": get_migration_status(db)
            }
        else:
            return {
                "message": "Some migrations failed. Check logs for details.",
                "migrations_run": migrations_run,
                "status": "failed"
            }
    except Exception as e:
        return {
            "message": f"Migration error: {str(e)}",
            "status": "error"
        }

@app.get("/health")
def health_check(db: Session = Depends(get_db)):
    """Health check endpoint for Render monitoring"""
    try:
        # Test database connection
        db.execute(text("SELECT 1"))

        # Check migration status
        migration_status = get_migration_status(db)

        return {
            "status": "healthy",
            "database": "connected",
            "migration": migration_status["migration_complete"],
            "timestamp": datetime.utcnow().isoformat()
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.utcnow().isoformat()
        }

# --- Routes ---

@app.post("/add_item")
def add_item(
    name: str,
    quantity: float,
    unit: str,
    price_per_unit: Optional[float] = None,
    total_cost: Optional[float] = None,
    type: Optional[str] = None,
    date_added: datetime = datetime.utcnow(),
    db: Session = Depends(get_db)
):

    if price_per_unit is not None:
        total_cost = quantity * price_per_unit
    elif total_cost is not None:
        total_cost = total_cost
    else:
        raise HTTPException(status_code=400, detail="Either price_per_unit or total_cost must be provided")


    item = Inventory(
        name=name,
        quantity=quantity,
        unit=unit,
        price_per_unit=price_per_unit if price_per_unit is not None else 0.0,
        total_cost=total_cost,
        type=type if type is not None else "",
        date_added=date_added
    )

    db.add(item)
    db.add(Expense(item_name=name, quantity=quantity, total_cost=total_cost, date=date_added))
    db.commit()
    db.refresh(item)

    inventory = db.query(Inventory).all()
    return {"message": "Item added successfully!", "inventory": inventory}

@app.get("/search_inventory")
def search_inventory(
    name: Optional[str] = Query(None, description="Partial or full item name"),
    type: Optional[str] = Query(None, description="Inventory type (e.g. 'Vegetables', 'Dairy')"),
    start_date: Optional[str] = Query(None, description="Start date in YYYY-MM-DD format"),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Inventory)

        if name:
            query = query.filter(Inventory.name.ilike(f"%{name}%"))
        elif type:
            query = query.filter(Inventory.type.ilike(f"%{type}%"))

        if start_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(Inventory.date_added >= start)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD.")

        if end_date:
            try:
                end = datetime.strptime(end_date, "%Y-%m-%d")
                query = query.filter(Inventory.date_added <= end)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD.")

        results = query.all()

        return [
            {
                "id": item.id,
                "name": item.name,
                "price_per_unit": item.price_per_unit,
                "unit": item.unit,
                "quantity": item.quantity,
                "total_cost": item.total_cost,
                "type": item.type,
                "date_added": item.date_added.strftime("%Y-%m-%d %H:%M:%S")
            }
            for item in results
        ]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")

@app.delete("/delete_item/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    db.delete(item)
    db.commit()
    return {
        "message": "Item deleted successfully!",
        "current_inventory": db.query(Inventory).all()
    }


@app.put("/update_item/{item_id}")
def update_item(
    item_id: int,
    name: str = Query(...),
    quantity: float = Query(...),
    unit: str = Query(...),
    price_per_unit: float = Query(None),
    total_cost: float = Query(None),
    date_added: Optional[datetime] = Query(default=datetime.utcnow()),
    type: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Recalculate total_cost
    if price_per_unit is not None:
        total_cost = quantity * price_per_unit
    elif total_cost is not None:
        total_cost = total_cost
    else:
        raise HTTPException(status_code=400, detail="Either price_per_unit or total_cost must be provided")

    # Update fields
    item.name = name
    item.quantity = quantity
    item.unit = unit
    item.price_per_unit = price_per_unit if price_per_unit is not None else 0.0
    item.total_cost = total_cost
    item.date_added = date_added
    item.type = type if type is not None else ""

    db.commit()
    db.refresh(item)

    return {
        "message": "Item updated successfully!",
        "updated_item": {
            "id": item.id,
            "name": item.name,
            "quantity": item.quantity,
            "unit": item.unit,
            "price_per_unit": item.price_per_unit,
            "total_cost": item.total_cost,
            "type": item.type,
            "date_added": item.date_added.strftime("%Y-%m-%d")
        },
        "current_inventory": [
            {
                "id": i.id,
                "name": i.name,
                "quantity": i.quantity,
                "unit": i.unit,
                "price_per_unit": i.price_per_unit,
                "total_cost": i.total_cost,
                "type": i.type,
                "date_added": i.date_added.strftime("%Y-%m-%d")
            } for i in db.query(Inventory).all()
        ]
    }

@app.get("/inventory")
def get_inventory(db: Session = Depends(get_db)):
    inventory_items = db.query(Inventory).all()
    return [
        {
            "id": item.id,
            "name": item.name,
            "price_per_unit": item.price_per_unit,
            "unit": item.unit,
            "quantity": item.quantity,
            "total_cost": item.total_cost,
            "type": item.type,
            "date_added": item.date_added.strftime("%Y-%m-%d %H:%M:%S")
        }
        for item in inventory_items
    ]

@app.get("/inventory_by_name/{item_name}")
def get_inventory_by_name(item_name: str, db: Session = Depends(get_db)):
    items = db.query(Inventory).filter(Inventory.name.ilike(f"%{item_name}%")).all()
    if not items:
        raise HTTPException(status_code=404, detail="Item not found")
    return items

@app.delete("/delete_all_inventory")
def delete_all_inventory(confirm: bool = False, db: Session = Depends(get_db)):
    if not confirm:
        raise HTTPException(status_code=400, detail="Please confirm deletion by setting confirm=true")

    deleted = db.query(Inventory).delete()
    db.commit()
    return {
        "message": f"Deleted {deleted} item(s) from inventory.",
        "current_inventory": db.query(Inventory).all()
    }


@app.get("/expense_report")
def expense_report(
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    inventory_name: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    # Determine date range if not provided
    if not start_date or not end_date:
        date_range = db.query(
            func.min(Inventory.date_added),
            func.max(Inventory.date_added)
        ).first()
        if not date_range or not date_range[0] or not date_range[1]:
            return {
                "message": "No inventory records in the database.",
                "total_expense": 0,
                "average_expense": 0,
                "highest_expense_day": None,
                "lowest_expense_day": None,
                "highest_expense_item": None,
                "lowest_expense_item": None,
                "most_frequent_inventory": None
            }
        start = date_range[0]
        end = date_range[1]
    else:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")

    # Base query
    query = db.query(Inventory).filter(
        Inventory.date_added >= start,
        Inventory.date_added <= end
    )

    # Apply inventory_name filter if present
    if inventory_name:
        query = query.filter(Inventory.name.ilike(f"%{inventory_name}%"))
    # Else apply type filter if present
    elif type:
        query = query.filter(Inventory.type.ilike(f"%{type}%"))

    inventory_items = query.all()

    if not inventory_items:
        return {
            "message": "No inventory found for the given filters.",
            "total_expense": 0,
            "average_expense": 0,
            "highest_expense_day": None,
            "lowest_expense_day": None,
            "highest_expense_item": None,
            "lowest_expense_item": None,
            "most_frequent_inventory": None
        }

    # Use total_cost directly
    total_expense = sum(item.total_cost for item in inventory_items)
    average_expense = total_expense / len(inventory_items)

    # Daily totals
    daily_expenses = defaultdict(float)
    for item in inventory_items:
        daily_expenses[item.date_added.date()] += item.total_cost

    highest_day = max(daily_expenses.items(), key=lambda x: x[1])
    lowest_day = min(daily_expenses.items(), key=lambda x: x[1])

    highest_expense_item = None
    lowest_expense_item = None
    most_frequent_inventory = None

    # Apply extended analysis only if inventory_name is NOT provided
    if not inventory_name:
        highest_item = max(inventory_items, key=lambda x: x.total_cost, default=None)
        lowest_item = min(inventory_items, key=lambda x: x.total_cost, default=None)
        highest_expense_item = highest_item.name if highest_item else None
        lowest_expense_item = lowest_item.name if lowest_item else None

        item_counts = Counter(item.name for item in inventory_items)
        most_frequent_inventory = item_counts.most_common(1)[0][0] if item_counts else None

    return {
        "inventory_name": inventory_name,
        "type": type if not inventory_name else None,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "total_expense": total_expense,
        "average_expense": average_expense,
        "highest_expense_day": {
            "date": highest_day[0].isoformat(),
            "amount": highest_day[1]
        },
        "lowest_expense_day": {
            "date": lowest_day[0].isoformat(),
            "amount": lowest_day[1]
        },
        "highest_expense_item": highest_expense_item,
        "lowest_expense_item": lowest_expense_item,
        "most_frequent_inventory": most_frequent_inventory
    }

@app.post("/upload_inventory_excel")
async def upload_inventory_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file (.xlsx or .xls)")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        required_columns = {"name", "quantity", "unit", "date_added"}
        optional_columns = {"price_per_unit", "total_cost", "type"}

        missing_required = required_columns - set(headers)
        if missing_required:
            return JSONResponse(status_code=400, content={
                "error": f"Missing required columns: {', '.join(missing_required)}. Found: {headers}"
            })

        col_index = {key: headers.index(key) for key in required_columns.union(optional_columns) if key in headers}

        added_items = []
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                name = row[col_index["name"]]
                quantity = float(row[col_index["quantity"]])
                unit = row[col_index["unit"]]

                price_per_unit = float(row[col_index["price_per_unit"]]) if "price_per_unit" in col_index and row[col_index["price_per_unit"]] is not None else None
                total_cost = float(row[col_index["total_cost"]]) if "total_cost" in col_index and row[col_index["total_cost"]] is not None else None
                type_ = str(row[col_index["type"]]).strip() if "type" in col_index and row[col_index["type"]] is not None else ""

                if price_per_unit is not None:
                    total_cost = quantity * price_per_unit
                elif total_cost is not None:
                    price_per_unit = total_cost / quantity
                else:
                    skipped_rows.append(f"Row {idx}: Missing price_per_unit and total_cost.")
                    continue

                date_raw = row[col_index["date_added"]]
                if isinstance(date_raw, datetime):
                    date_added = date_raw
                elif isinstance(date_raw, str):
                    date_added = datetime.strptime(date_raw.strip(), "%Y-%m-%d")
                else:
                    raise ValueError("Invalid date format in 'date_added'")

                # Check if an exact item already exists
                existing = db.query(Inventory).filter(
                    Inventory.name == name,
                    Inventory.quantity == quantity,
                    Inventory.unit == unit,
                    Inventory.price_per_unit == price_per_unit,
                    Inventory.total_cost == total_cost,
                    Inventory.type == type_,
                    Inventory.date_added == date_added
                ).first()

                if existing:
                    skipped_rows.append(f"Row {idx}: Exact item already exists. Skipped.")
                    continue

                item = Inventory(
                    name=name,
                    quantity=quantity,
                    unit=unit,
                    price_per_unit=price_per_unit,
                    total_cost=total_cost,
                    type=type_,
                    date_added=date_added
                )
                db.add(item)
                db.add(Expense(item_name=name, quantity=quantity, total_cost=total_cost, date=date_added))
                added_items.append(name)

            except Exception as row_error:
                skipped_rows.append(f"Row {idx}: {str(row_error)}")

        db.commit()

        return {
            "message": "Excel processed successfully",
            "added_items": added_items,
            "skipped_rows": skipped_rows
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@app.post("/upload_dish_excel")
async def upload_dish_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file (.xlsx or .xls)")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        required_columns = {"name", "type", "ingredient_name", "quantity_required"}

        missing_required = required_columns - set(headers)
        if missing_required:
            return JSONResponse(status_code=400, content={
                "error": f"Missing required columns: {', '.join(missing_required)}. Found: {headers}"
            })

        col_index = {key: headers.index(key) for key in required_columns}

        dishes_map = {}  # Temporary cache to hold dishes before commit
        added_dishes = []
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                dish_name = str(row[col_index["name"]]).strip()
                dish_type = str(row[col_index["type"]]).strip()
                ingredient_name = str(row[col_index["ingredient_name"]]).strip()
                quantity_required = float(row[col_index["quantity_required"]])

                # Fetch or create DishType
                dish_type_obj = db.query(DishType).filter_by(name=dish_type).first()
                if not dish_type_obj:
                    dish_type_obj = DishType(name=dish_type)
                    db.add(dish_type_obj)
                    db.flush()  # To assign an ID

                # Unique key for dish mapping
                dish_key = (dish_name, dish_type_obj.id)
                if dish_key not in dishes_map:
                    dish = db.query(Dish).filter_by(name=dish_name, type_id=dish_type_obj.id).first()
                    if not dish:
                        dish = Dish(name=dish_name, type_id=dish_type_obj.id)
                        db.add(dish)
                        db.flush()  # Assign ID
                        added_dishes.append(dish_name)
                    dishes_map[dish_key] = dish
                else:
                    dish = dishes_map[dish_key]

                # Add ingredient
                dish_ingredient = DishIngredient(
                    dish_id=dish.id,
                    ingredient_name=ingredient_name,
                    quantity_required=quantity_required
                )
                db.add(dish_ingredient)

            except Exception as row_error:
                skipped_rows.append(f"Row {idx}: {str(row_error)}")

        db.commit()

        return {
            "message": "Dishes uploaded successfully",
            "added_dishes": list(set(added_dishes)),
            "skipped_rows": skipped_rows
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@app.post("/add_dish")
def add_dish(request: AddDishRequest, db: Session = Depends(get_db)):
    # Check if dish with same name exists
    existing_dish = db.query(Dish).filter(Dish.name == request.name).first()
    if existing_dish:
        raise HTTPException(status_code=400, detail="Dish with this name already exists.")

    # Get or create DishType
    dish_type = db.query(DishType).filter(DishType.name == request.type).first()
    if not dish_type:
        dish_type = DishType(name=request.type)
        db.add(dish_type)
        db.commit()
        db.refresh(dish_type)

    # Create new dish
    new_dish = Dish(name=request.name, type_id=dish_type.id)
    db.add(new_dish)
    db.commit()
    db.refresh(new_dish)

    # Process ingredients
    for ing in request.ingredients:
        #inventory_item = db.query(Inventory).filter(Inventory.name == ing.name).first()
        #if not inventory_item:
            #raise HTTPException(status_code=404, detail=f"Ingredient '{ing.name}' not found in inventory.")

        dish_ingredient = DishIngredient(
            dish_id=new_dish.id,
            #ingredient_name=inventory_item.name,
            ingredient_name=ing.name,
            quantity_required=ing.quantity_required
        )
        db.add(dish_ingredient)

    db.commit()
    return {"message": f"Dish '{request.name}' added successfully with ingredients."}

@app.get("/dishes", response_model=List[DishOut])
def list_dishes(db: Session = Depends(get_db)):
    dishes = db.query(Dish).all()
    result = []

    for dish in dishes:
        # Get the dish type name
        dish_type = db.query(DishType).filter(DishType.id == dish.type_id).first()

        # Get all ingredients for this dish
        ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

        # Convert DishIngredient DB entries to response models
        ingredient_list = [
            DishIngredientOut(
                ingredient_name=di.ingredient_name,
                quantity_required=di.quantity_required
            ) for di in ingredients if di.ingredient_name is not None
        ]

        # Append formatted dish to result
        result.append(DishOut(
            id=dish.id,
            name=dish.name,
            type=dish_type.name if dish_type else "Unknown",
            ingredients=ingredient_list
        ))

    return result


@app.get("/dishes/by_name", response_model=List[DishOut])
def search_dishes_by_name(partial_name: str, db: Session = Depends(get_db)):
    matched_dishes = db.query(Dish).filter(
        Dish.name.ilike(f"%{partial_name}%")
    ).all()

    if not matched_dishes:
        raise HTTPException(status_code=404, detail="No matching dishes found")

    result = []
    for dish in matched_dishes:
        dish_type = db.query(DishType).filter(DishType.id == dish.type_id).first()
        ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

        ingredient_list = [
            DishIngredientOut(
                ingredient_name=di.ingredient_name,
                quantity_required=di.quantity_required
            )
            for di in ingredients
        ]

        result.append(DishOut(
            id=dish.id,
            name=dish.name,
            type=dish_type.name if dish_type else "Unknown",
            ingredients=ingredient_list
        ))

    return result

@app.delete("/dishes/{dish_name}")
def delete_dish_by_name(
    dish_name: str,
    confirm: bool = Query(False, description="Set to true to confirm deletion"),
    db: Session = Depends(get_db)
):
    dish = db.query(Dish).filter(Dish.name.ilike(dish_name)).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    if not confirm:
        raise HTTPException(
            status_code=400,
            detail=f"Deletion not confirmed. To delete dish '{dish.name}', set confirm=true."
        )

    db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).delete()
    db.delete(dish)
    db.commit()
    return {"message": f"Dish '{dish.name}' deleted successfully"}

@app.get("/dish_types")
def get_dish_types(db: Session = Depends(get_db)):
    dish_types = db.query(DishType).all()
    return [dt.name for dt in dish_types]


@app.get("/dishes/{dish_id}/cost")
def get_dish_cost(dish_id: int, db: Session = Depends(get_db)):
    dish = db.query(Dish).filter(Dish.id == dish_id).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()
    total_cost = 0.0
    ingredient_costs = []

    for di in ingredients:
        # Get the most recent inventory item for this ingredient
        inventory_item = db.query(Inventory).filter(
            Inventory.name.ilike(di.ingredient_name)
        ).order_by(Inventory.date_added.desc()).first()

        if not inventory_item:
            raise HTTPException(
                status_code=400,
                detail=f"Ingredient '{di.ingredient_name}' not found in inventory"
            )

        item_cost = di.quantity_required * inventory_item.price_per_unit
        total_cost += item_cost

        ingredient_costs.append({
            "ingredient": di.ingredient_name,
            "quantity_required": di.quantity_required,
            "unit_price": inventory_item.price_per_unit,
            "total_cost": item_cost
        })

    return {
        "dish_id": dish.id,
        "dish_name": dish.name,
        "ingredient_breakdown": ingredient_costs,
        "total_cost": round(total_cost, 2)
    }


@app.put("/dishes/{dish_id}")
def update_dish(dish_id: int, payload: DishUpdate, db: Session = Depends(get_db)):
    dish = db.query(Dish).filter(Dish.id == dish_id).first()
    if not dish:
        raise HTTPException(status_code=404, detail="Dish not found")

    # Update dish name and type
    dish.name = payload.name

    # Lookup dish type
    dish_type = db.query(DishType).filter(DishType.name.ilike(payload.type)).first()
    if not dish_type:
        dish_type = DishType(name=payload.type)
        db.add(dish_type)
        db.commit()
        db.refresh(dish_type)

    dish.type_id = dish_type.id
    db.commit()

    dish_ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()

    # Update ingredients: map by name for easy comparison
    existing_ingredients = {di.ingredient_name.lower(): di for di in dish_ingredients}
    updated_names = {ing.ingredient_name.lower() for ing in payload.ingredients}

    for ing in payload.ingredients:
        key = ing.ingredient_name.lower()
        if key in existing_ingredients:
            existing_ingredients[key].quantity_required = ing.quantity_required
        else:
            db.add(DishIngredient(
                dish_id=dish.id,
                ingredient_name=ing.ingredient_name,
                quantity_required=ing.quantity_required
            ))

    # Delete ingredients no longer in request

    for key, di in existing_ingredients.items():
        if key not in updated_names:
            db.delete(di)

    db.commit()
    return {"message": "Dish updated successfully"}


@app.post("/prepare_dish")
def prepare_dish(
    dish_name: str = Query(..., description="Name of the dish to prepare"),
    quantity: float = Query(..., description="Number of servings"),
    date: Optional[str] = Query(None, description="Preparation date in YYYY-MM-DD format"),
    db: Session = Depends(get_db)
):
    # Parse preparation date
    if date:
        try:
            prepare_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    else:
        prepare_date = datetime.utcnow()

    # Fetch dish
    dish = db.query(Dish).filter(Dish.name.ilike(dish_name.strip())).first()
    if not dish:
        raise HTTPException(status_code=404, detail=f"Dish '{dish_name}' not found")

    # Get ingredients
    dish_ingredients = db.query(DishIngredient).filter(DishIngredient.dish_id == dish.id).all()
    if not dish_ingredients:
        raise HTTPException(status_code=400, detail=f"No ingredients found for dish '{dish_name}'")

    usage_summary = []

    for ingredient in dish_ingredients:
        required_qty = ingredient.quantity_required * quantity

        inventory_batches = db.query(Inventory).filter(
            Inventory.name.ilike(ingredient.ingredient_name),
            Inventory.quantity > 0
        ).order_by(Inventory.date_added.asc()).all()

        if not inventory_batches:
            raise HTTPException(
                status_code=400,
                detail=f"No inventory available for {ingredient.ingredient_name}"
            )

        for batch in inventory_batches:
            if required_qty <= 0:
                break

            unit = batch.unit.strip().lower()
            batch_qty_in_base = batch.quantity

            # Unit conversion
            if unit == "kg":
                batch_qty_in_base *= 1000
            elif unit in ["g", "gm", "grams"]:
                pass
            elif unit in ["liter", "litre", "l"]:
                batch_qty_in_base *= 1000
            elif unit == "ml":
                pass
            elif unit in ["piece", "pieces", "pc", "pcs", "pack", "packs"]:
                pass
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported unit '{batch.unit}' for ingredient '{batch.name}'")

            # Deduct from this batch
            deduct_qty = min(batch_qty_in_base, required_qty)
            remaining_qty_in_base = batch_qty_in_base - deduct_qty
            required_qty -= deduct_qty

            # Convert back to original unit
            if unit == "kg":
                batch.quantity = remaining_qty_in_base / 1000
            elif unit in ["g", "gm", "grams"]:
                batch.quantity = remaining_qty_in_base
            elif unit in ["liter", "litre", "l"]:
                batch.quantity = remaining_qty_in_base / 1000
            elif unit == "ml":
                batch.quantity = remaining_qty_in_base
            elif unit in ["piece", "pieces", "pc", "pcs", "pack", "packs"]:
                batch.quantity = remaining_qty_in_base

            db.add(batch)

            # Insert log for this batch
            db.add(InventoryLog(
                ingredient_id=batch.id,
                quantity_left=batch.quantity,
                date=prepare_date
            ))

            usage_summary.append({
                "inventory_name": ingredient.ingredient_name,
                "inventory_batch_id": batch.id,
                "used_from_batch": deduct_qty,
                "remaining_in_batch": batch.quantity,
                "unit": unit,
                "logged_at": prepare_date.strftime("%Y-%m-%d")
            })

            # Update all future logs of this batch
            future_logs = db.query(InventoryLog).filter(
                and_(
                    InventoryLog.ingredient_id == batch.id,
                    InventoryLog.date > prepare_date
                )
            ).order_by(InventoryLog.date.asc()).all()

            current_quantity = batch.quantity

            for log in future_logs:
                log.quantity_left = current_quantity
                db.add(log)

        # ❗ Still required quantity means not enough inventory
        if required_qty > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Not enough '{ingredient.ingredient_name}' in inventory to prepare {quantity} servings. Short by {required_qty:.2f} units."
            )

        # ✅ Recalculate inventory quantity across ALL batches of this ingredient
        all_batches = db.query(Inventory).filter(
            Inventory.name.ilike(ingredient.ingredient_name)
        ).all()

        for b in all_batches:
            latest_log = db.query(InventoryLog).filter(
                InventoryLog.ingredient_id == b.id
            ).order_by(InventoryLog.date.desc()).first()

            if latest_log:
                b.quantity = latest_log.quantity_left
                db.add(b)

    db.commit()

    return {
        "message": f"Dish '{dish.name}' prepared successfully for {quantity} servings on {prepare_date.date()}.",
        "usage_summary": usage_summary
    }


@app.post("/upload_prepare_dish_excel")
async def upload_prepare_dish_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload a .xlsx or .xls file.")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
        required_columns = {"dish_name", "quantity", "date"}
        col_index = {h: i for i, h in enumerate(headers)}

        missing = required_columns - set(col_index.keys())
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

        success_count = 0
        failed_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                dish_name = str(row[col_index["dish_name"]]).strip()
                quantity = float(row[col_index["quantity"]])
                #date_str = str(row[col_index["date"]]).strip()
                date_str = row[col_index["date"]]
                date = datetime.strptime(date_str, "%Y-%m-%d")

                #date = parser.parse(date_str).date()
                '''
                date_raw = row[col_index["date"]]
                if isinstance(date_raw, datetime):
                    date = date_raw
                elif isinstance(date_raw, str):
                    date= datetime.strptime(date_raw.strip(), "%Y-%m-%d")
                else:
                    raise ValueError("Invalid date format in 'date_added'")
                '''
                # Call your existing dish preparation logic here
                # This is a simplified version
                response = prepare_dish(dish_name, quantity, date, db)
                if response.get("success"):
                    success_count += 1
                else:
                    failed_rows.append(f"Row {idx}: {response.get('error')}")

            except Exception as e:
                failed_rows.append(f"Row {idx}: {str(e)}")

        return {
            "message": f"Processed {success_count} row(s) successfully.",
            "errors": failed_rows
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@app.get("/inventory_on_date")
def inventory_on_date(date: str, db: Session = Depends(get_db)):
    try:
        date_parsed = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    # Step 1: Get all unique ingredient_ids from logs
    ingredient_ids = db.query(InventoryLog.ingredient_id).distinct().all()
    ingredient_ids = [i[0] for i in ingredient_ids]

    response = []

    # Step 2: For each ingredient, get the latest log <= selected date
    for ingredient_id in ingredient_ids:
        latest_log = db.query(InventoryLog).filter(
            InventoryLog.ingredient_id == ingredient_id,
            InventoryLog.date <= date_parsed
        ).order_by(InventoryLog.date.desc()).first()

        if latest_log:
            inventory_item = db.query(Inventory).filter(Inventory.id == ingredient_id).first()
            response.append({
                "ingredient_id": ingredient_id,
                "ingredient_name": inventory_item.name if inventory_item else "Unknown",
                "unit": inventory_item.unit if inventory_item else "",
                "quantity_left": latest_log.quantity_left,
                "log_time": latest_log.date.strftime("%Y-%m-%d %H:%M:%S")
            })

    return response


class OpenAIPromptRequest(BaseModel):
    prompt: str

@app.post("/ask_openai")
def ask_openai(request: OpenAIPromptRequest, db: Session = Depends(get_db)):
    try:
        prompt = request.prompt.strip()

        # Limit Inventory and Expense records
        #inventory_items = db.query(Inventory).order_by(Inventory.date_added.desc()).limit(50).all()
        inventory_items = db.query(Inventory).order_by(Inventory.date_added.desc()).all()
        #expenses = db.query(Expense).order_by(Expense.date.desc()).limit(50).all()

        # Summarized Inventory Summary including Total Cost
        inventory_summary = "\n".join(
            [
                f"{item.name} ({item.quantity} {item.unit}), Type: {item.type}, Total Cost: ₹{item.total_cost}"
                for item in inventory_items
            ]
        )

        '''
        # Summarized Expense Summary
        expense_summary = "\n".join(
            [
                f"{exp.item_name}: ₹{exp.total_cost} on {exp.date.date()}"
                for exp in expenses
            ]
        )
        '''

        # System prompt
        system_prompt = (
            "You are an expert restaurant Inventory and Expense report generator."
            " Use the provided data below to answer the user's specific question or create a report."
            "\n\nInventory Data (latest 50 items):\n" + inventory_summary
        )

        # OpenAI Request
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo-1106",  # or gpt-4-1106-preview if you have access
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=1500,   # Safe
        )

        ai_message = response['choices'][0]['message']['content']
        return {"response": ai_message}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"OpenAI Error: {str(e)}")

@app.get("/debug/database-info")
def get_database_info():
    return {
        "database_url": DATABASE_URL.split('@')[0] + '@***' if '@' in DATABASE_URL else DATABASE_URL,
        "database_type": "postgresql" if "postgresql" in DATABASE_URL else "sqlite",
        "engine": str(engine.url).split('@')[0] + '@***' if '@' in str(engine.url) else str(engine.url)
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
